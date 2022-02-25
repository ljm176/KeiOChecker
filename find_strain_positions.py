# -*- coding: utf-8 -*-
"""
Created on Thu Feb 17 14:45:25 2022

@author: lajamu
"""

import openpyxl



layoutwb = openpyxl.load_workbook("Plates.xlsx")
layoutSheet = layoutwb.active

def find_position(gene):
    for row in range(2, layoutSheet.max_row):
        for col in range(3, layoutSheet.max_column+1):
            if layoutSheet.cell(row=row, column = col).value == gene:
                return [layoutSheet.cell(row=row, column = 1).value,
                        str(layoutSheet.cell(row=row, column = 2).value) +
                        str(layoutSheet.cell(row=1, column = col).value)]
                            
                        
def list_positions(genes_list):
    l = []
    for gene in genes_list:
        l.append(find_position(gene))
    return(l)
