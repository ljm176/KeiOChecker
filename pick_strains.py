# -*- coding: utf-8 -*-
"""
Created on Thu Feb 17 14:45:25 2022

@author: lajamu
"""

import openpyxl

strainswb = openpyxl.load_workbook("keioPrimers1.xlsx")
strainsSheet = strainswb.active
genes = []
for row in range(2, strainsSheet.max_row +1):
        genes.append(strainsSheet["A" + str(row)].value)


layoutwb = openpyxl.load_workbook("Plates.xlsx")
layoutSheet = layoutwb.active

def find_position(gene):
    for row in range(2, layoutSheet.max_row):
        for col in range(3, layoutSheet.max_column+1):
            if layoutSheet.cell(row=row, column = col).value == gene:
                return [layoutSheet.cell(row=row, column = 1).value,
                        str(layoutSheet.cell(row=row, column = 2).value) +
                        str(layoutSheet.cell(row=1, column = col).value)]
                            
                        

