# -*- coding: utf-8 -*-
"""
Created on Tue Feb  8 10:26:54 2022

@author: lajamu
"""

from Bio import SeqIO
from Bio.Seq import Seq
import openpyxl
from Bio.SeqUtils import MeltingTemp as mt

coliGenome = SeqIO.read("MG1655.fasta", "fasta")
coli_seq = coliGenome.seq 

wb = openpyxl.load_workbook("msb4100050-s4.xlsx")
sheet=  wb.active

gene_dict = {}

for g in range(4, sheet.max_row):
    gene = sheet["B" + str(g)].value
    seq1 = sheet["K" + str(g)].value
    seq2 = sheet["L" + str(g)].value

    gene_dict[gene] = seq1

#TODO: Add feature to lengthen or shorten based on Tm
def get_primer_loc_5prime(gene):
    target = Seq(gene_dict[gene])
    target_loc = coli_seq.find(target)
    primer_loc = target_loc - 500
    primer = coli_seq[primer_loc - 20 : primer_loc]
    n = 20
    count=0

    while not 62 < mt.Tm_Wallace(primer) < 66 and not count > 20:
        print(gene)
        print(mt.Tm_Wallace(primer))
        print("Count: " + str(count) )
        print(n)
        print(len(primer))
        
        if mt.Tm_Wallace(primer) > 66:
            primer = coli_seq[primer_loc - n+1 : primer_loc]
            n -=1
        if mt.Tm_Wallace(primer) < 62:
            primer = coli_seq[primer_loc - n-1 : primer_loc]
            n +=1 
        count += 1
    return(primer)
    
def get_primer_loc_3prime(gene):
    target = Seq(gene_dict[gene]).reverse_complement()
    target_loc = coli_seq.find(target)
    primer_loc = target_loc + 500
    primer = coli_seq[primer_loc - 20 : primer_loc].reverse_complement()
    n = 20
    count = 0
    while not 62 < mt.Tm_Wallace(primer) < 66 and not count > 20:

        if mt.Tm_Wallace(primer) > 66:
            primer = coli_seq[primer_loc - n-1 : primer_loc].reverse_complement()
            n -=1
        if mt.Tm_Wallace(primer) < 62:
            primer = coli_seq[primer_loc - n+1 : primer_loc].reverse_complement()
            n +=1 
        count += 1
    return(primer)
    

def get_primer(gene):
    if coli_seq.find(gene_dict[gene]) == -1:
        primer = get_primer_loc_3prime(gene)
        if coli_seq.count(primer.reverse_complement()) !=  1:
            return("Double binding site, manually design this primer")
        else:
            return(primer)
    else:
        primer = get_primer_loc_5prime(gene)
        if coli_seq.count(primer) != 1:
            return("Double binding site, manually design this primer")
        else:
            return(primer)
    return(primer)
        
def write_primer_to_xl(f):
    wb2 = openpyxl.load_workbook(f)
    sheet2 = wb2.active
    genes = []
    
    for row in range(2, sheet2.max_row +1):
        genes.append(sheet2["A" + str(row)].value)
    print(genes)
    primers = []
    for g in genes:
        primers.append(get_primer(g))
    sheet2["B1"].value = "Primers"
    sheet2["C1"].value = "Tm"
    print(primers)
    
    for p, r in zip(primers, range(2, sheet2.max_row +1)):
        sheet2["B" + str(r)].value = str(p)
        sheet2["C" + str(r)].value = mt.Tm_Wallace(p)
    primerFile = f[:-5] + "_Primers.xlsx"
    wb2.save(primerFile)
    
    return(primerFile)
    
write_primer_to_xl("keioPrimers1.xlsx")
    
    
    
